from __future__ import annotations

import io
from datetime import datetime
from functools import wraps
from pathlib import Path

from flask import (
    Flask,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, text
from sqlalchemy.exc import IntegrityError
from werkzeug.security import check_password_hash, generate_password_hash

BASE_DIR = Path(__file__).resolve().parent

app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{BASE_DIR / 'attendance.db'}"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SECRET_KEY"] = "lab-attendance-secret-key"

db = SQLAlchemy(app)


class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)


class SignInRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_name = db.Column(db.String(100), nullable=False)
    computer_name = db.Column(db.String(150), nullable=False)
    signed_at = db.Column(db.DateTime, nullable=False, default=datetime.now)


class SignInConfig(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    start_time = db.Column(db.DateTime, nullable=True)
    end_time = db.Column(db.DateTime, nullable=True)


class TeacherAccount(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    password_hash = db.Column(db.String(255), nullable=False)


SEAT_LAYOUT = [
    [60, 59, None, None, None, None, 16, 15],
    [58, 57, 44, 43, 30, 29, 14, 13],
    [56, 55, 42, 41, 28, 27, 12, 11],
    [54, 53, 40, 39, 26, 25, 10, 9],
    [52, 51, 38, 37, 24, 23, 8, 7],
    [50, 49, 36, 35, 22, 21, 6, 5],
    [48, 47, 34, 33, 20, 19, 4, 3],
    [46, 45, 32, 31, 18, 17, 2, 1],
]


def is_teacher_logged_in() -> bool:
    return bool(session.get("teacher_logged_in"))


def teacher_required(func_view):
    @wraps(func_view)
    def wrapper(*args, **kwargs):
        if is_teacher_logged_in():
            return func_view(*args, **kwargs)

        if request.path.startswith("/api/"):
            return jsonify({"ok": False, "message": "请先登录教师端。"}), 401
        return redirect(url_for("teacher_login_page"))

    return wrapper


def init_db() -> None:
    db.create_all()
    if not db.session.get(SignInConfig, 1):
        db.session.add(SignInConfig(id=1))

    teacher = db.session.get(TeacherAccount, 1)
    if not teacher:
        db.session.add(TeacherAccount(id=1, password_hash=generate_password_hash("abc123")))

    deduped: set[str] = set()
    for record in SignInRecord.query.order_by(SignInRecord.signed_at.asc(), SignInRecord.id.asc()).all():
        key = record.student_name.lower()
        if key in deduped:
            db.session.delete(record)
        else:
            deduped.add(key)

    db.session.commit()
    db.session.execute(
        text(
            "CREATE UNIQUE INDEX IF NOT EXISTS idx_unique_signin_student "
            "ON sign_in_record(lower(student_name))"
        )
    )
    db.session.commit()


def resolve_client_name() -> str:
    forwarded = request.headers.get("X-Forwarded-For")
    ip = (forwarded.split(",")[0].strip() if forwarded else request.remote_addr) or "unknown"
    return ip


def parse_dt(value: str | None) -> datetime | None:
    if not value:
        return None
    return datetime.strptime(value, "%Y-%m-%dT%H:%M")


def get_status_payload() -> dict:
    students = [s.name for s in Student.query.order_by(Student.name).all()]
    records = SignInRecord.query.order_by(SignInRecord.signed_at.desc()).all()
    signed_info = {r.student_name.lower(): r for r in records}
    cfg = db.session.get(SignInConfig, 1)

    roster = []
    for name in students:
        row = signed_info.get(name.lower())
        roster.append(
            {
                "student_name": name,
                "status": "已签到" if row else "未签到",
                "computer_name": row.computer_name if row else "-",
                "signed_at": row.signed_at.strftime("%Y-%m-%d %H:%M:%S") if row else "-",
                "signed_at_dt": row.signed_at if row else None,
            }
        )

    roster.sort(
        key=lambda item: (
            0 if item["signed_at_dt"] is not None else 1,
            item["signed_at_dt"] if item["signed_at_dt"] is not None else datetime.max,
            item["student_name"],
        )
    )
    for item in roster:
        item.pop("signed_at_dt", None)

    return {
        "students": students,
        "roster": roster,
        "records": [
            {
                "student_name": r.student_name,
                "computer_name": r.computer_name,
                "signed_at": r.signed_at.strftime("%Y-%m-%d %H:%M:%S"),
            }
            for r in records
        ],
        "signed_count": len(signed_info),
        "total_count": len(students),
        "absent_count": max(len(students) - len(signed_info), 0),
        "window": {
            "start": cfg.start_time.strftime("%Y-%m-%d %H:%M") if cfg and cfg.start_time else None,
            "end": cfg.end_time.strftime("%Y-%m-%d %H:%M") if cfg and cfg.end_time else None,
        },
    }



def get_seat_grid() -> list[list[dict]]:
    records = SignInRecord.query.order_by(SignInRecord.signed_at.asc()).all()
    seat_to_names: dict[int, list[str]] = {}
    for r in records:
        parts = (r.computer_name or "").split(".")
        if len(parts) != 4 or not parts[-1].isdigit():
            continue
        seat_no = int(parts[-1])
        if not 1 <= seat_no <= 60:
            continue
        seat_to_names.setdefault(seat_no, []).append(r.student_name)

    grid: list[list[dict]] = []
    for row in SEAT_LAYOUT:
        out_row = []
        for seat_no in row:
            if seat_no is None:
                out_row.append({"seat_no": None, "label": "", "students": []})
            else:
                names = seat_to_names.get(seat_no, [])
                out_row.append({
                    "seat_no": seat_no,
                    "label": str(seat_no),
                    "students": names,
                })
        grid.append(out_row)
    return grid


@app.get("/")
def index():
    return redirect(url_for("student_page"))


@app.get("/student")
def student_page():
    return render_template("student.html", computer_name=resolve_client_name())


@app.get("/teacher/login")
def teacher_login_page():
    if is_teacher_logged_in():
        return redirect(url_for("teacher_page"))
    return render_template("teacher_login.html")


@app.post("/teacher/login")
def teacher_login_api():
    password = (request.form.get("password") or "").strip()
    account = db.session.get(TeacherAccount, 1)

    if not account or not check_password_hash(account.password_hash, password):
        return render_template("teacher_login.html", error="密码错误，请重试。"), 401

    session["teacher_logged_in"] = True
    return redirect(url_for("teacher_page"))


@app.post("/teacher/logout")
def teacher_logout_api():
    session.pop("teacher_logged_in", None)
    return redirect(url_for("teacher_login_page"))


@app.get("/teacher")
@teacher_required
def teacher_page():
    return render_template("teacher.html")


@app.get("/teacher/seats")
@teacher_required
def teacher_seats_page():
    return render_template("seat_map.html", seat_grid=get_seat_grid())


@app.get("/api/students")
def students_api():
    q = request.args.get("q", "").strip().lower()
    query = Student.query.order_by(Student.name)
    if q:
        query = query.filter(func.lower(Student.name).contains(q))
    names = [s.name for s in query.limit(15).all()]
    return jsonify({"students": names})


@app.get("/api/status")
@teacher_required
def status_api():
    return jsonify(get_status_payload())


@app.post("/api/signin")
def signin_api():
    payload = request.get_json(force=True)
    student_name = (payload.get("student_name") or "").strip()

    if not student_name:
        return jsonify({"ok": False, "message": "请输入姓名。"}), 400

    student = Student.query.filter(func.lower(Student.name) == student_name.lower()).first()
    if not student:
        return jsonify({"ok": False, "message": "该姓名不在名单中，请联系老师。"}), 400

    cfg = db.session.get(SignInConfig, 1)
    now = datetime.now()
    if cfg and cfg.start_time and now < cfg.start_time:
        return jsonify({"ok": False, "message": "签到未开始，请在规定时间内签到。"}), 400
    if cfg and cfg.end_time and now > cfg.end_time:
        return jsonify({"ok": False, "message": "签到时间已结束。"}), 400

    existing = SignInRecord.query.filter(func.lower(SignInRecord.student_name) == student.name.lower()).first()
    if existing:
        return jsonify({"ok": False, "message": "你已签到，无需重复提交。"}), 400

    db.session.add(
        SignInRecord(
            student_name=student.name,
            computer_name=resolve_client_name(),
            signed_at=now,
        )
    )
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify({"ok": False, "message": "你已签到，无需重复提交。"}), 400
    return jsonify({"ok": True, "message": f"{student.name} 签到成功！"})


@app.post("/api/reset")
@teacher_required
def reset_api():
    SignInRecord.query.delete()
    db.session.commit()
    return jsonify({"ok": True, "message": "签到记录已重置。"})


@app.post("/api/clear-roster")
@teacher_required
def clear_roster_api():
    SignInRecord.query.delete()
    Student.query.delete()
    db.session.commit()
    return jsonify({"ok": True, "message": "当前名单与签到记录已清空。"})


@app.post("/api/change-password")
@teacher_required
def change_password_api():
    payload = request.get_json(force=True)
    old_password = (payload.get("old_password") or "").strip()
    new_password = (payload.get("new_password") or "").strip()

    if len(new_password) < 6:
        return jsonify({"ok": False, "message": "新密码长度至少 6 位。"}), 400

    account = db.session.get(TeacherAccount, 1)
    if not account or not check_password_hash(account.password_hash, old_password):
        return jsonify({"ok": False, "message": "旧密码不正确。"}), 400

    account.password_hash = generate_password_hash(new_password)
    db.session.commit()
    return jsonify({"ok": True, "message": "密码修改成功。"})


@app.post("/api/window")
@teacher_required
def set_window_api():
    payload = request.get_json(force=True)
    cfg = db.session.get(SignInConfig, 1)
    cfg.start_time = parse_dt(payload.get("start_time"))
    cfg.end_time = parse_dt(payload.get("end_time"))
    db.session.commit()
    return jsonify({"ok": True, "message": "签到时间段已更新。"})


@app.post("/api/import")
@teacher_required
def import_students_api():
    file = request.files.get("file")
    if not file:
        return jsonify({"ok": False, "message": "请上传 Excel 文件。"}), 400

    wb = load_workbook(file, data_only=True)
    ws = wb.active
    added = 0
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        cell = row[0]
        if not cell:
            continue
        name = str(cell).strip()
        if not name:
            continue
        exists = Student.query.filter(func.lower(Student.name) == name.lower()).first()
        if exists:
            continue
        db.session.add(Student(name=name))
        added += 1
    db.session.commit()
    return jsonify({"ok": True, "message": f"导入完成，新增 {added} 名学生。"})


@app.get("/api/export")
@teacher_required
def export_records_api():
    wb = Workbook()
    ws = wb.active
    ws.title = "签到记录"
    ws.append(["姓名", "签到状态", "计算机IP", "签到时间"])

    signed_map: dict[str, SignInRecord] = {}
    for record in SignInRecord.query.order_by(SignInRecord.signed_at.desc()).all():
        signed_map[record.student_name.lower()] = record

    for student in Student.query.order_by(Student.name).all():
        matched = signed_map.get(student.name.lower())
        ws.append(
            [
                student.name,
                "已签到" if matched else "未签到",
                matched.computer_name if matched else "",
                matched.signed_at.strftime("%Y-%m-%d %H:%M:%S") if matched else "",
            ]
        )

    buff = io.BytesIO()
    wb.save(buff)
    buff.seek(0)
    return send_file(
        buff,
        as_attachment=True,
        download_name=f"signin_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    with app.app_context():
        init_db()
    app.run(host="0.0.0.0", port=5000, debug=True)
