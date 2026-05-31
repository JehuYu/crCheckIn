from __future__ import annotations

import io
from itertools import combinations
import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE_COLUMNS = ["学号", "考号", "姓名", "班级", "学校", "全卷", "1卷", "2卷"]
TOTAL_COLUMNS = ["全卷", "1卷", "2卷", "信息", "通用"]
DISPLAY_TOTAL_COLUMNS = ["全卷", "信息", "通用"]
CLASS_ASSIGNMENT_FILE = "高一教学班汇总表（202604）.xlsx"
SUBJECT_INFO = "信息"
SUBJECT_GENERAL = "通用"
SUBJECT_UNKNOWN = "未分类"


@dataclass
class QuestionColumn:
    name: str
    column_index: int
    subject: str
    has_answer: bool = False


@dataclass
class ExamSchema:
    base_columns: list[str]
    question_columns: list[QuestionColumn]
    answer_columns: list[str] = field(default_factory=list)
    excluded_subtotal_columns: list[str] = field(default_factory=list)
    unclassified_columns: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    class_group_columns: list[str] = field(default_factory=list)
    unmatched_students: dict[str, list[str]] = field(default_factory=dict)


@dataclass
class AnalysisResult:
    data: pd.DataFrame
    schema: ExamSchema
    class_averages: pd.DataFrame
    info_analysis: pd.DataFrame
    general_analysis: pd.DataFrame

    @property
    def summary(self) -> dict:
        info_count = sum(1 for q in self.schema.question_columns if q.subject == SUBJECT_INFO)
        general_count = sum(1 for q in self.schema.question_columns if q.subject == SUBJECT_GENERAL)
        unknown_count = sum(1 for q in self.schema.question_columns if q.subject == SUBJECT_UNKNOWN)
        return {
            "question_count": len(self.schema.question_columns),
            "info_question_count": info_count,
            "general_question_count": general_count,
            "unclassified_question_count": unknown_count,
            "excluded_subtotal_columns": self.schema.excluded_subtotal_columns,
            "unclassified_columns": self.schema.unclassified_columns,
            "warnings": self.schema.warnings,
        }


def _clean_header(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def _make_unique(names: Iterable[str]) -> list[str]:
    seen: dict[str, int] = {}
    result = []
    for raw_name in names:
        name = raw_name or "未命名列"
        count = seen.get(name, 0)
        seen[name] = count + 1
        result.append(name if count == 0 else f"{name}_{count + 1}")
    return result


def _normalize_question_name(name: str) -> str:
    text = str(name).strip()
    text = text.replace("_", "-").replace("－", "-").replace("—", "-")
    text = re.sub(r"\s+", "", text)
    return text


def natural_sort_key(name: str) -> tuple:
    text = _normalize_question_name(name)
    if "草图" in text:
        tail = 999
    else:
        tail = 0
    subject_rank = 0 if text.startswith("信") else 1 if text.startswith("通") or text.startswith("T") else 2
    numbers = [int(n) for n in re.findall(r"\d+", text)]
    return (subject_rank, numbers, tail, text)


def infer_subject(question_name: str) -> str:
    text = _normalize_question_name(question_name)
    if text.startswith("信"):
        return SUBJECT_INFO
    if text.startswith("通") or text.startswith("T"):
        return SUBJECT_GENERAL

    first_number = re.search(r"\d+", text)
    if first_number:
        number = int(first_number.group(0))
        if 13 <= number <= 15:
            return SUBJECT_INFO
        if 28 <= number <= 30:
            return SUBJECT_GENERAL
    return SUBJECT_UNKNOWN


def read_exam_file(file_obj) -> pd.DataFrame:
    """Read the new two-row-header score workbook as raw cell data."""
    try:
        raw = pd.read_excel(file_obj, header=None, dtype=object)
    except ImportError as exc:
        raise ValueError("缺少读取 Excel 所需依赖，请安装 pandas、openpyxl 和 xlrd 后重试。") from exc
    except Exception as exc:
        raise ValueError(f"无法读取 Excel 文件：{exc}") from exc

    if raw.shape[0] < 3:
        raise ValueError("成绩表至少需要包含两行表头和一行学生数据。")
    if raw.shape[1] < len(BASE_COLUMNS):
        raise ValueError("成绩表列数不足，未找到完整的学生基础字段。")
    return raw


def detect_exam_schema(raw: pd.DataFrame, mapping_file: str = CLASS_ASSIGNMENT_FILE) -> tuple[pd.DataFrame, ExamSchema]:
    """Detect base fields, score columns, answer columns and subject grouping."""
    header_top = [_clean_header(v) for v in raw.iloc[0].tolist()]
    header_sub = [_clean_header(v) for v in raw.iloc[1].tolist()]

    base_headers = header_top[: len(BASE_COLUMNS)]
    missing_base = [name for name in BASE_COLUMNS if name not in base_headers]
    if missing_base:
        raise ValueError(f"缺少关键字段：{', '.join(missing_base)}。请确认使用新格式成绩表。")

    output_names: list[str] = []
    score_names: list[str] = []
    answer_names: list[str] = []
    raw_questions: list[QuestionColumn] = []
    last_question_name = ""

    for index, top_name in enumerate(header_top):
        sub_name = header_sub[index] if index < len(header_sub) else ""
        if index < len(BASE_COLUMNS):
            output_names.append(top_name)
            continue

        if top_name:
            last_question_name = top_name

        if sub_name == "答案":
            question_name = last_question_name or top_name or f"第{index + 1}列"
            col_name = f"{question_name}_答案"
            answer_names.append(col_name)
            output_names.append(col_name)
            continue

        if sub_name == "得分" or top_name:
            question_name = top_name or last_question_name or f"第{index + 1}列"
            subject = infer_subject(question_name)
            score_names.append(question_name)
            raw_questions.append(QuestionColumn(question_name, index, subject, has_answer=(sub_name == "得分")))
            output_names.append(question_name)
            continue

        output_names.append(f"未命名列{index + 1}")

    output_names = _make_unique(output_names)
    data = raw.iloc[2:].copy()
    data.columns = output_names
    data = data.dropna(subset=["学号"]).reset_index(drop=True)

    for col in BASE_COLUMNS:
        if col in data.columns:
            data[col] = data[col].astype(str).str.strip()
    data["班级"] = data["班级"].map(normalize_source_class)

    for col in ["全卷", "1卷", "2卷"] + score_names:
        if col in data.columns:
            data[col] = pd.to_numeric(data[col], errors="coerce")

    questions = _sync_question_names(raw_questions, output_names)
    excluded_subtotals = _detect_subtotal_columns(data, questions)
    questions = [q for q in questions if q.name not in excluded_subtotals]
    questions = sorted(questions, key=lambda q: natural_sort_key(q.name))

    unclassified = [q.name for q in questions if q.subject == SUBJECT_UNKNOWN]
    warnings = []
    if excluded_subtotals:
        warnings.append(f"已自动排除疑似汇总列：{', '.join(excluded_subtotals)}")
    if unclassified:
        warnings.append(f"以下题目未能判断信息/通用归属：{', '.join(unclassified)}")
    if not questions:
        raise ValueError("未识别到有效小题得分列。")

    data, class_group_columns, class_warnings, unmatched_students = attach_teaching_classes(data, mapping_file)
    warnings.extend(class_warnings)

    schema = ExamSchema(
        base_columns=BASE_COLUMNS.copy(),
        question_columns=questions,
        answer_columns=answer_names,
        excluded_subtotal_columns=excluded_subtotals,
        unclassified_columns=unclassified,
        warnings=warnings,
        class_group_columns=class_group_columns,
        unmatched_students=unmatched_students,
    )
    return data, schema


def attach_teaching_classes(
    data: pd.DataFrame, mapping_file: str = CLASS_ASSIGNMENT_FILE
) -> tuple[pd.DataFrame, list[str], list[str], dict[str, list[str]]]:
    warnings = []
    unmatched_students: dict[str, list[str]] = {}
    if not pd.io.common.file_exists(mapping_file):
        warnings.append(f"未找到分班表：{mapping_file}，只生成原始班级明细。")
        return data, [], warnings, unmatched_students

    mapping = pd.read_excel(mapping_file, dtype=object)
    required_columns = {"班级", "姓名"}
    if not required_columns.issubset(mapping.columns):
        warnings.append("分班表缺少“班级”或“姓名”字段，无法生成教学班明细。")
        return data, [], warnings, unmatched_students

    teaching_columns = [col for col in mapping.columns if str(col).startswith("教学班")]
    if not teaching_columns:
        warnings.append("分班表未找到“教学班”字段，无法生成教学班明细。")
        return data, [], warnings, unmatched_students

    mapping = mapping.copy()
    mapping["班级"] = mapping["班级"].astype(str).str.strip()
    mapping["姓名"] = mapping["姓名"].astype(str).str.strip()
    mapping["_匹配班级"] = mapping["班级"].map(_normalize_class_for_join)

    rename_map = {}
    for index, col in enumerate(teaching_columns, start=1):
        rename_map[col] = "信息教学班" if index == 1 else "通用教学班" if index == 2 else f"教学班{index}"
    mapping = mapping.rename(columns=rename_map)
    class_group_columns = list(rename_map.values())

    merged = data.copy()
    merged["班级"] = merged["班级"].astype(str).str.strip()
    merged["姓名"] = merged["姓名"].astype(str).str.strip()
    merged["_匹配班级"] = merged["班级"].map(_normalize_class_for_join)
    merged = merged.merge(
        mapping[["_匹配班级", "姓名"] + class_group_columns],
        on=["_匹配班级", "姓名"],
        how="left",
    )
    merged = merged.drop(columns=["_匹配班级"])

    for col in class_group_columns:
        missing_mask = merged[col].isna() & (merged["班级"] != "台州分校")
        missing_count = int(missing_mask.sum())
        if missing_count:
            students = [
                f"{row['班级']} {row['姓名']}"
                for _, row in merged.loc[missing_mask, ["班级", "姓名"]].iterrows()
            ]
            unmatched_students[col] = students
            warnings.append(f"{col} 有 {missing_count} 名学生未在分班表中匹配到：{', '.join(students)}")

    return merged, class_group_columns, warnings, unmatched_students


def _normalize_class_for_join(class_name: str) -> str:
    text = str(class_name).strip()
    match = re.fullmatch(r"一(\d+)", text)
    if match:
        return match.group(1)
    return text


def normalize_source_class(class_name: str) -> str:
    text = str(class_name).strip()
    if re.fullmatch(r"一\d+", text):
        return "台州分校"
    return text


def _sync_question_names(questions: list[QuestionColumn], output_names: list[str]) -> list[QuestionColumn]:
    synced = []
    for question in questions:
        col_name = output_names[question.column_index]
        synced.append(
            QuestionColumn(
                name=col_name,
                column_index=question.column_index,
                subject=question.subject,
                has_answer=question.has_answer,
            )
        )
    return synced


def _detect_subtotal_columns(data: pd.DataFrame, questions: list[QuestionColumn]) -> list[str]:
    score_names = [q.name for q in questions if q.name in data.columns]
    normalized = {name: _normalize_question_name(name) for name in score_names}
    excluded = []

    for candidate in score_names:
        candidate_norm = normalized[candidate]
        parts = [
            name
            for name in score_names
            if name != candidate and re.match(rf"^{re.escape(candidate_norm)}[-_]\d+$", normalized[name])
        ]
        if len(parts) < 2:
            continue

        candidate_values = pd.to_numeric(data[candidate], errors="coerce").fillna(0)
        matching_parts = _find_matching_subtotal_parts(data, candidate_values, parts)
        if matching_parts:
            excluded.append(candidate)

    return sorted(excluded, key=natural_sort_key)


def _find_matching_subtotal_parts(data: pd.DataFrame, candidate_values: pd.Series, parts: list[str]) -> tuple[str, ...]:
    numeric_parts = data[parts].apply(pd.to_numeric, errors="coerce").fillna(0)
    max_size = min(len(parts), 6)
    for size in range(2, max_size + 1):
        for selected in combinations(parts, size):
            selected_sum = numeric_parts[list(selected)].sum(axis=1)
            if (candidate_values - selected_sum).abs().max() <= 0.001:
                return selected
    return ()


def analyze_exam_scores(data: pd.DataFrame, schema: ExamSchema) -> AnalysisResult:
    score_columns = [q.name for q in schema.question_columns]
    info_cols = [q.name for q in schema.question_columns if q.subject == SUBJECT_INFO]
    general_cols = _sort_general_detail_columns(
        [q.name for q in schema.question_columns if q.subject == SUBJECT_GENERAL]
    )

    prepared = data.copy()
    prepared["信息"] = prepared[info_cols].sum(axis=1).round(2) if info_cols else 0
    prepared["通用"] = prepared[general_cols].sum(axis=1).round(2) if general_cols else 0

    class_averages = _build_group_average(prepared, DISPLAY_TOTAL_COLUMNS + score_columns)
    info_analysis = _build_group_average(prepared, ["信息"] + info_cols)
    general_analysis = _build_group_average(prepared, ["通用"] + general_cols)

    return AnalysisResult(
        data=prepared,
        schema=schema,
        class_averages=class_averages,
        info_analysis=info_analysis,
        general_analysis=general_analysis,
    )


def _build_group_average(data: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    available = [col for col in columns if col in data.columns]
    if not available:
        return pd.DataFrame()

    class_avg = data.groupby("班级")[available].mean(numeric_only=True).round(2)
    class_avg.index = class_avg.index.astype(str)
    class_avg = class_avg.sort_index()

    total_avg = data[available].mean(numeric_only=True).round(2)
    total_avg = pd.DataFrame([total_avg], index=["总平均值"])
    return pd.concat([total_avg, class_avg])


def process_exam_file(file_obj, mapping_file: str = CLASS_ASSIGNMENT_FILE) -> AnalysisResult:
    raw = read_exam_file(file_obj)
    data, schema = detect_exam_schema(raw, mapping_file)
    return analyze_exam_scores(data, schema)


def read_excel_file(file_obj):
    """Compatibility wrapper for existing callers."""
    raw = read_exam_file(file_obj)
    data, _schema = detect_exam_schema(raw)
    return data


def classify_by_subject(data):
    info_cols = [col for col in data.columns if infer_subject(col) == SUBJECT_INFO and col not in BASE_COLUMNS + TOTAL_COLUMNS]
    general_cols = [col for col in data.columns if infer_subject(col) == SUBJECT_GENERAL and col not in BASE_COLUMNS + TOTAL_COLUMNS]
    return sorted(info_cols, key=natural_sort_key), sorted(general_cols, key=natural_sort_key)


def calculate_class_averages(data):
    info_cols, general_cols = classify_by_subject(data)
    score_cols = sorted(set(info_cols + general_cols), key=natural_sort_key)
    prepared = data.copy()
    if "信息" not in prepared:
        prepared["信息"] = prepared[info_cols].sum(axis=1).round(2) if info_cols else 0
    if "通用" not in prepared:
        prepared["通用"] = prepared[general_cols].sum(axis=1).round(2) if general_cols else 0
    return _build_group_average(prepared, DISPLAY_TOTAL_COLUMNS + score_cols)


def generate_info_analysis(data, info_cols):
    prepared = data.copy()
    if "信息" not in prepared:
        prepared["信息"] = prepared[info_cols].sum(axis=1).round(2) if info_cols else 0
    return _build_group_average(prepared, ["信息"] + list(info_cols))


def generate_general_analysis(data, general_cols):
    prepared = data.copy()
    if "通用" not in prepared:
        prepared["通用"] = prepared[general_cols].sum(axis=1).round(2) if general_cols else 0
    return _build_group_average(prepared, ["通用"] + list(general_cols))


def add_color_scale(ws, start_col, end_col, start_row, end_row):
    if end_row < start_row:
        return
    rule = ColorScaleRule(
        start_type="min",
        start_color="FFF16B61",
        mid_type="percentile",
        mid_value=50,
        mid_color="FFF7E98F",
        end_type="max",
        end_color="FF64BC7B",
    )
    ws.conditional_formatting.add(f"{start_col}{start_row}:{end_col}{end_row}", rule)


def save_results_to_excel(all_avg, info_analysis, general_analysis, output_file, data, schema=None):
    wb = build_excel_report(
        AnalysisResult(
            data=data,
            schema=schema or _schema_from_data(data),
            class_averages=all_avg,
            info_analysis=info_analysis,
            general_analysis=general_analysis,
        )
    )
    wb.save(output_file)


def save_results_to_excel_bytes(all_avg, info_analysis, general_analysis, data, schema=None):
    wb = build_excel_report(
        AnalysisResult(
            data=data,
            schema=schema or _schema_from_data(data),
            class_averages=all_avg,
            info_analysis=info_analysis,
            general_analysis=general_analysis,
        )
    )
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def save_analysis_to_excel(result: AnalysisResult, output_file):
    wb = build_excel_report(result)
    wb.save(output_file)


def save_analysis_to_excel_bytes(result: AnalysisResult):
    output = io.BytesIO()
    build_excel_report(result).save(output)
    output.seek(0)
    return output


def _schema_from_data(data: pd.DataFrame) -> ExamSchema:
    question_columns = []
    for col in data.columns:
        if col in BASE_COLUMNS + TOTAL_COLUMNS or col.endswith("_答案"):
            continue
        subject = infer_subject(col)
        if subject != SUBJECT_UNKNOWN:
            question_columns.append(QuestionColumn(col, -1, subject))
    return ExamSchema(BASE_COLUMNS.copy(), sorted(question_columns, key=lambda q: natural_sort_key(q.name)))


def build_excel_report(result: AnalysisResult) -> Workbook:
    wb = Workbook()
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws = wb.active
    ws.title = "班级平均分"
    _write_dataframe_sheet(ws, result.class_averages, "班级", border, with_color_scale=True)
    _write_dataframe_sheet(wb.create_sheet("信息学科分析"), result.info_analysis, "班级", border, with_color_scale=True)
    _write_dataframe_sheet(wb.create_sheet("通用学科分析"), result.general_analysis, "班级", border, with_color_scale=True)

    _write_warning_sheet(wb, result, border)
    for group_column in result.schema.class_group_columns:
        subject = SUBJECT_INFO if group_column == "信息教学班" else SUBJECT_GENERAL if group_column == "通用教学班" else None
        prefix = "一信" if group_column == "信息教学班" else "一通" if group_column == "通用教学班" else group_column
        _write_class_detail_sheets(wb, result, border, group_column, prefix, subject=subject)
    _write_class_detail_sheets(wb, result, border, "班级", "")
    return wb


def _write_dataframe_sheet(ws, df: pd.DataFrame, index_name: str, border: Border, with_color_scale=False):
    ws.append([index_name] + list(df.columns))
    for idx, row in df.iterrows():
        ws.append([idx] + [_excel_value(v) for v in row.values])
    _style_used_range(ws, border)

    if with_color_scale and df.shape[0] > 0 and df.shape[1] > 0:
        for col_idx in range(2, df.shape[1] + 2):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            add_color_scale(ws, col_letter, col_letter, 2, df.shape[0] + 1)


def _write_warning_sheet(wb: Workbook, result: AnalysisResult, border: Border):
    ws = wb.create_sheet("识别摘要")
    summary_rows = [
        ["识别项目", "结果"],
        ["有效小题数", result.summary["question_count"]],
        ["信息题数", result.summary["info_question_count"]],
        ["通用题数", result.summary["general_question_count"]],
        ["未分类题数", result.summary["unclassified_question_count"]],
        ["排除汇总列", "、".join(result.schema.excluded_subtotal_columns) or "无"],
        ["未分类题目", "、".join(result.schema.unclassified_columns) or "无"],
    ]
    for row in summary_rows:
        ws.append(row)

    ws.append([])
    ws.append(["提示"])
    for warning in result.schema.warnings or ["无"]:
        ws.append([warning])

    ws.append([])
    ws.append(["自动排除的汇总列"])
    ws.append(["列名"])
    for column_name in result.schema.excluded_subtotal_columns or ["无"]:
        ws.append([column_name])

    ws.append([])
    ws.append(["未匹配分班表学生"])
    ws.append(["教学班类型", "学生"])
    if result.schema.unmatched_students:
        for group_column, students in result.schema.unmatched_students.items():
            for student in students:
                ws.append([group_column, student])
    else:
        ws.append(["无", "无"])
    _style_used_range(ws, border)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 80


def _write_class_detail_sheets(
    wb: Workbook,
    result: AnalysisResult,
    border: Border,
    group_column: str,
    sheet_prefix: str,
    subject: str | None = None,
):
    data = result.data.copy()
    if group_column not in data.columns:
        return

    info_cols = [q.name for q in result.schema.question_columns if q.subject == SUBJECT_INFO]
    general_cols = _sort_general_detail_columns(
        [q.name for q in result.schema.question_columns if q.subject == SUBJECT_GENERAL]
    )
    unknown_cols = [q.name for q in result.schema.question_columns if q.subject == SUBJECT_UNKNOWN]
    if subject == SUBJECT_INFO:
        general_cols = []
        unknown_cols = []
        base_columns = ["姓名", "班级", group_column, "全卷", "信息", "通用"]
    elif subject == SUBJECT_GENERAL:
        info_cols = []
        unknown_cols = []
        base_columns = ["姓名", "班级", group_column, "全卷", "信息", "通用"]
    else:
        base_columns = ["姓名", "班级", "全卷", "信息", "通用"]
    if group_column != "班级":
        if group_column not in base_columns:
            base_columns.insert(2, group_column)

    for class_name in sorted(data[group_column].dropna().astype(str).unique()):
        class_data = data[data[group_column].astype(str) == class_name]
        title = _teaching_sheet_title(sheet_prefix, class_name) if sheet_prefix else class_name
        ws = wb.create_sheet(_safe_sheet_title(title, wb.sheetnames))
        info_columns = base_columns + info_cols + unknown_cols
        general_columns = base_columns + general_cols

        if subject == SUBJECT_INFO:
            _write_detail_table(
                ws,
                _sort_detail_data(class_data, ["信息", "全卷"]),
                info_columns,
                start_row=1,
                border=border,
            )
        elif subject == SUBJECT_GENERAL:
            _write_detail_table(
                ws,
                _sort_detail_data(class_data, ["通用", "全卷"]),
                general_columns,
                start_row=1,
                border=border,
            )
        else:
            next_row = _write_detail_table(
                ws,
                _sort_detail_data(class_data, ["信息"]),
                info_columns,
                start_row=1,
                border=border,
            )
            _write_detail_table(
                ws,
                _sort_detail_data(class_data, ["通用"]),
                general_columns,
                start_row=next_row + 2,
                border=border,
            )
        _fit_sheet(ws)


def _sort_detail_data(data: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    sort_columns = [col for col in columns if col in data.columns]
    if not sort_columns:
        return data
    return data.sort_values(sort_columns, ascending=[False] * len(sort_columns), na_position="last")


def _sort_general_detail_columns(columns: list[str]) -> list[str]:
    normal = [col for col in columns if "补线" not in col and "草图" not in col]
    repair = [col for col in columns if "补线" in col]
    sketch = [col for col in columns if "草图" in col]
    return (
        sorted(normal, key=natural_sort_key)
        + sorted(repair, key=natural_sort_key)
        + sorted(sketch, key=natural_sort_key)
    )


def _write_detail_table(ws, class_data: pd.DataFrame, columns: list[str], start_row: int, border: Border) -> int:
    header_fill = PatternFill(start_color="FFCCCCCC", end_color="FFCCCCCC", fill_type="solid")
    zero_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")

    for col_index, column_name in enumerate(columns, start=1):
        cell = ws.cell(row=start_row, column=col_index, value=column_name)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_offset, (_, row) in enumerate(class_data.iterrows(), start=1):
        excel_row = start_row + row_offset
        for col_index, column_name in enumerate(columns, start=1):
            value = _excel_value(row.get(column_name, ""))
            cell = ws.cell(row=excel_row, column=col_index, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if col_index > 5 and value == 0:
                cell.fill = zero_fill

    return start_row + len(class_data) + 1


def _style_used_range(ws, border: Border):
    header_fill = PatternFill(start_color="FFCCCCCC", end_color="FFCCCCCC", fill_type="solid")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = Font(bold=True)
        ws.row_dimensions[row[0].row].height = 16

    for column in ws.columns:
        column_letter = column[0].column_letter
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
        ws.column_dimensions[column_letter].width = max(8, min(max_len + 2, 24))

def _fit_sheet(ws):
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 16

    for column in ws.columns:
        column_letter = column[0].column_letter
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
        ws.column_dimensions[column_letter].width = max(8, min(max_len + 2, 24))


def _safe_sheet_title(title: str, existing: list[str]) -> str:
    cleaned = re.sub(r"[\[\]\*:/\\?]", "_", str(title)).strip() or "班级"
    cleaned = cleaned[:31]
    candidate = cleaned
    counter = 2
    while candidate in existing:
        suffix = f"_{counter}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        counter += 1
    return candidate


def _teaching_sheet_title(prefix: str, class_name: str) -> str:
    text = str(class_name).strip()
    match = re.search(r"([AB]\d+)$", text)
    if match:
        return f"{prefix}{match.group(1)}"
    return f"{prefix}{text}"


def _excel_value(value):
    if pd.isna(value):
        return None
    return value


def main():
    result = process_exam_file("data/小题分(技术选).xls")
    output_file = f"分析结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    save_analysis_to_excel(result, output_file)
    print(f"分析结果已保存到 {output_file}")


if __name__ == "__main__":
    main()
